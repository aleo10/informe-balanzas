"""
Script de una sola vez para crear base_datos.xlsx con datos de ejemplo.
Ejecutar: python crear_base_datos.py
Luego editar el Excel para agregar todos los clientes y balanzas reales.
"""
import openpyxl
import os

OUT = os.path.join(os.path.dirname(__file__), "data", "base_datos.xlsx")
os.makedirs(os.path.dirname(OUT), exist_ok=True)

wb = openpyxl.Workbook()

# ── Hoja Clientes ─────────────────────────────────────────────────────────────
ws_cli = wb.active
ws_cli.title = "Clientes"
ws_cli.append(["Razon Social", "Direccion", "Localidad"])
clientes = [
    ["SALTA REFRESCOS S.A.", "INGENIO BELLA VISTA", "BELLA VISTA - TUCUMAN"],
    ["TEMAS INDUSTRIALES S.A.", "RN 34 KM 1085", "ROSARIO DE LA FRONTERA - SALTA"],
]
for row in clientes:
    ws_cli.append(row)

# ── Hoja Balanzas ─────────────────────────────────────────────────────────────
ws_bal = wb.create_sheet("Balanzas")
headers = [
    "Razon Social",
    "Tipo Balanza",
    "Mod Indicador",
    "Marca",
    "N Serie Indicador",
    "Cod Aprobacion Indicador",
    "Ubicacion",
    "Plataforma",
    "Medidas",
    "N Serie Plataforma",
    "Cod Aprobacion Plataforma",
    "Ult Verificacion",
    "Cap Max kg",
    "Cap Min kg",
    "Div Min kg",
    "Apoyos",
    "Cod Interno",
    "Precintos Anteriores Plataforma",
    "Precintos Vigentes Plataforma",
    "Precintos Anteriores Indicador",
    "Precintos Vigentes Indicador",
]
ws_bal.append(headers)

balanzas = [
    ["SALTA REFRESCOS S.A.", "CAMIONES", "LE 100 1", "LATORRE",
     "311009", "BF.60-1933", "ENTRADA", "CAMIONES", "30mts",
     "311009", "BF.80-2083", "21/8/2024",
     100000, 1000, 20, 8, "PRODUC. TERMI.",
     "", "R083 - R085 - R086 - R087",
     "", "R173 - R174 - R081 - R082 - R084"],

    ["SALTA REFRESCOS S.A.", "Embolsadora", "Onix", "Sipel",
     "43914", "--------", "SALA DE EMBOLSADO", "TOLVA", "",
     "43914", "--------", "",
     100, "", 0.05, "", "Embolsadora 01",
     "", "", "", ""],

    ["SALTA REFRESCOS S.A.", "SPC", "Onix", "Sipel",
     "48084", "--------", "JUGO", "TOLVA", "",
     "48084", "--------", "",
     15000, "", 5, "", "JUGO1",
     "", "", "", ""],

    ["SALTA REFRESCOS S.A.", "Embolsadora Big Bag", "Onix", "Sipel",
     "43920", "--------", "PLANTA", "TOLVA", "",
     "43920", "--------", "",
     2000, "", 0.5, "", "BIG BAG",
     "", "", "", ""],

    ["SALTA REFRESCOS S.A.", "SPC", "Onix", "Sipel",
     "48100", "--------", "MELAZA", "TOLVA", "",
     "48100", "--------", "",
     15000, "", 5, "", "MELAZA",
     "", "", "", ""],

    ["TEMAS INDUSTRIALES S.A.", "Balanza de piso", "ABC", "XYZ",
     "12345", "--------", "PLANTA", "PLATAFORMA", "",
     "12345", "--------", "",
     500, "", 0.2, "", "PISO-01",
     "", "", "", ""],
]
for row in balanzas:
    ws_bal.append(row)

# Ajustar anchos
for ws in [ws_cli, ws_bal]:
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

# ── Hoja Configuracion ────────────────────────────────────────────────────────
ws_cfg = wb.create_sheet("Configuracion")
ws_cfg.append(["Clave", "Valor"])
ws_cfg.append(["cert_pesas_pequenas",  "00769-S-1225"])
ws_cfg.append(["desc_pesas_pequenas",  "1kg, 2kg, 5kg, 10kg, 20kg"])
ws_cfg.append(["cert_pesas_grandes",   "00763-S-1225"])
ws_cfg.append(["desc_pesas_grandes",   "500kg, 1000kg"])
ws_cfg.column_dimensions["A"].width = 30
ws_cfg.column_dimensions["B"].width = 30

wb.save(OUT)
print(f"Base de datos creada en: {OUT}")
print("Abrila en Excel para agregar todos los clientes y balanzas.")
